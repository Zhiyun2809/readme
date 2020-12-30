#===============================
# Colab include mount my drive 
from google.colab import drive
drive.mount('/content/dirve')
#===============================
# get object info
for attr in dir(workObj):
    print(attr +'()') if callable(getattr(wo1,attr)) else print(attr)

#===============================
# check if string contains substring
if 'test' in stra:
    print(stra)
#===============================
#method for scaling
from sklearn import preprocessing
X = np.array(df)
X = preprocessing.scale(X)
#alternative method
from sklern.processing import StandardScaler
X = StandardScaler().fit_transform(select_df)

score= accuracy_score(data_true = data_test, data_pred=predictions)

# generate linear space depends on dataframe
x = pd.DataFrame('lstat':np.linspace(df['lstat'].min(),df['lstat'].max(),100))

# 
from sklearn.metrics import mean_squared_error
from math import sqrt
RMSE = sqrt(mean_squared_error(y_true=y_test, y_pred=y_prediction)

#===============================
#Pandas dataframe
#===============================
# expand mxinum range of rows to view
pd.set_option('display.max_rows',3000)
pd.set_option('display.max_columns',3000)
pd.set_option('display.width',3000)

#===============================
# displace numerical feature
pd.describe()
pd.describe().transpose()

#  get numerical colums
newdf = df.select_dtypes(include=numeric)
df.select_dtypes('number').columns

# 
titles = list(df.select_dtype(inlcude='category')
titles = list(df.select_dtype(exclude='category')

# get columns or row index
train_df.columns
train_df.index
train_df.columns.values

# convert column to digits
df['m_hct']=df['m_hct'].round(1) -> convert to 1 decimal place

# displace categorical feature (no numerical values)
pd.describe(include=['O'])
pd.describe(include='all')

pd.describe().T

# get 10 instances from df. 
df.sample(10)
df.sample(100,replace=True) #allow duplication

# get index of mininum column value
df['printtime'].idxmin(axis=)

train_df[['Pclass','Survived']].groupby(['Pclass'],as_index=False).mean().sort_values(by='Survived',ascending=False)
train_df[['log','red_move']].groupby(['log'],as_index=False).agg(['mean','count']).sort_values(by=('red_move','mean'),ascending=False)

df['smoker'].value_counts()

# display value frequency
train_df['smoker'].value_counts()
train_df['smoker'].value_counts(normalize=True)

df1.groupby(['district','year']).count()['count']

# cal frequency count
df['country'].value_counts()
#===============================
import calendar
df['month'] = df['month'].apply(lambda x: calendar.monthabbr[x])

#===============================
# cleanup rare title names
stat_min = 10
title_names = (data1['Title'].value_counts() < stat_min)
data1['Title'] = data1['Title'].apply(lambda x:'Misc' if title_names.loc[x] == True else x)
combine = [train_df, test_df]
for dataset in combine:
    dataset['Title'] = dataset.Name.str.extract(' ([A-Za-z]+)\.', expand=False)
    pd.crosstab(train_df['Title'], train_df['Sex'])

display(pd.crosstab(df['DEATH_EVENT'],[df['smoking'],
                                        df['high_blood_pressure'],
                                        df['sex']],
                                        dropna=False))

# combine 2 columns
df['join'] = df['a']+df['b']
df['join'] = df[['a','b']].agg('_'.join, axis=1)
# apply
import re
def get_title(name):
    title_search = re.search(' (A-Za-z)+)\.'), name)
    if title_serach:
        return title_serach.group(1)
    return ''
dataset['Title'] = dataset['Name'].apply(get_title)

# convering categorial to number
title_mapping = {'Mr':1, 'Miss':2, 'Mrs':3, 'Master':4, 'Rare':5}
for dataset in combine:
    dataset['Title'] = dataset['Title'].map(title_mapping)
    dataset['Title'] = dataset['Title'].fillna(0)
# replace
recipes = recipes.replace(to_replace='Yes',value=1)
data = data.replace(-1,np.nan)
#mapping
for dataset in combine:
    dataset['Sex'] = dataset['Sex'].map({'female':1, 'male':0}).astype(int)
#fillna
df['Age'].fillna(method='ffill')    # forward fill
df['Age'].fillna(method='bfill')    # backward fill
#or
df = df.ffill(axis=0)       

# check is any nan in the dataframe
df.isnull().any() -> return list of False or True 
df.isnull().any().any() -> return False or True

df[df['A'].isnulll()].index
# alternative
# get overview of missing data
df.isnull().sum()
# get overview of missing data
df.isnull().sum().sum()
# list nan
df.loc[pd.isna(df['date']),:]
#  shift col
df[ticker].shift(-i)
# find common index
idx=df1.index.intersection(df2.index)
#
age_avg = dataset['Age'].mean()
age_std = dataset['Age'].std()
age_null_count = dataset['Age'].isnull().sum()
age_null_random_list = np.random.randint(age_avg-age_std, age_avg+age_std, size=age_null_count)
dataset['Age'][np.isnan(dataset['Age'])]=age_null_random_list
dataset['Age'] = dataset['Age'].astype(int)

# transfer dataframe to dict
grps = df['group'].unique().tolist()
d_data = {grp:data['weight'][data['group']==grp] for grp in grps}
part_grp = data.groupby('group').size()[0]

dict_vote = {'vote':np.where(np.random.rand(n) < 0.1,'Brown','Green')}
df_vote = pd.DataFrame({'vote':np.where(np.random.rand(n) < 0.1,'Brown','Green')})

# create dataframe from series
c = pd.DataFrame({'mean':df_mean,'std':df_std})

# transpose
df_t = df.transpose()

#filter row by list of strings
cleaned = df[~df['stn'].isin(remove_list)]
# 
df.filter(regex='e$',axis=0) # select row name ending with 'e'
df.filter(regex='^c',axis=0) # select row name starting with 'c'

# remove any instance(s) which has 2 digits followed by 5 digits and again 5 digits
df = df[df['Name'].str.contains(r'\d{2} \d{5} \d{5}')==False]

df = df.reset_index(drop = True)
# rename columns
df.rename(columns={"old_name":"new_name"},inplace=True)

# rolling operation
df['rolling_mean']=df['close'].rolling(window=3).mean()
# get the max indx idxmax
df.idxmax(axis=0)
df['sodium'].idxmax()

# handling redundacy
df.drop_duplicates()

# deprecated: df.loc[list-of-labels]
df.reindex(list-of-labels)

df.loc[df.index.intersection(labels)]
# fine top 3 words 
all_chat_list=[]
for i in range(len(df['Name'].drop_duplicates())):
    temp= df['Convo'][df['Name']==df['Name'].drop_duplicates().reset_index(drop=True)[i]]
    temp= temp.reset_index(drop = True)
    for j in range(1,len(temp)):
        temp[0] +=' '+temp[j]
    all_chat_list.append(temp[0])
    del temp
# select column
# get list of seceltion
df[i'libido'][df['dose']=='high']

# cut vs qcut
train_df['AgeBand'] = pd.cut(train_df['Age'], 5)
train_df[['AgeBand','Survived']].groupby(['AgeBand'], as_index=False).mean().sort_values(by='AgeBand', ascending=True )
smoking['ageGroup'] = pd.cut(smoking['age'],[0,30,40,53,54],labels['0-30','30-40','40-53','53-60'])
dataset['FareBin']=pd.qcut(dataset['Fare'],4)
dataset['AgeBin'] =pd.cut(dataset['Age'].astype(int),5)

# replace
df['dose'].replace({1:'placebo',2:'low'}, inplace=True)

df['pos']=df['settings'].str.replace(r'_10x_settings.txt','')

string.replace(old,new,count)

stat_min = 10
title_names = (data1['Title'].value_counts()< stat_min)



# lambda x:
data1['Title'] = data1['Title'].apply(lambda x: 'Misc' if title_names.loc[x]==True else x)
train['Has_Cabin'] = train['Cabin'].apply(lambda x: 0 if type(x)==float else 1)

item0 = lambda o:o[0]

df['Tax %'] = df['Income'].apply(lambda x: .15 if 10000<x<40000 else .2 if 40000<x<80000 else .25)
df ['Taxes owed']= df['Income'] * df['Tax %']

# filter value
df['Test Col'] = False
df.loc[df['Income']<6000,'Test Col'] = True
# loc
dataset.loc[dataset['Age']<=16, 'Age'] = 0
dataset.loc['IsAlone'].loc[dataset['FamilySize']>1] = 0
# loc
dataset.loc[dataset['FamilizeSize']==1, 'IsAlone'] = 1
dataset['Fare'] = dataset['Fare'].fillna(train['Fare'].median())

# map
df['Sex'] = df['Sex'].map({'female':0,'male':1}).astype(int)
# replace
df['dose'].replace({1:'placebo',2:'low'}, inplace=True)

# get the most frequent status
freq_port = train_df['Embarked'].dropna().mode()[0]
for dataset in combine:
    dataset['Embarked'] = dataset['Embarked'].fillna(freq_port)
# sort
df.sort_values(by=['Age'],ascending=False,inplace=True)
# set value
for index, row in data2.iterrows():
    data1.set_value(index,'radom',1)

# use pd.values
df = pd.DataFrame({col1: [2,3],col2:[3,4]})
dict_df = {}
for key, item in df.values:
    dict_df[key] = item


# dropna
df.dropna(axis=1)  # drop all columns with NaN
df.dropna(axis=0)  # drop rows iwth na
df.dropna(how='all') # drop rows where all elements are missing
df.dropna(thresh=2)  # drop rows with at least 2 NA vales
df.dropna(subset=['name'],axis=0) # look for NA only in subset columns
# drop name
df.drop(['Name','PassengerId'],axis=1)

# split column into multiple columns
new = df['pos'].str.split('_',n=1,expand=True)
df[['new1','new2']]= df['pos'].str.split('_',n=1,expand=True)
df['new1'],_ = df['pos'].str(split('\n',1).str
# strip
data['Min_Salary'],data['Max_Salary'] = data['Salary Estimate'].str.split('-').str
data['Min_Salary']=data['Min_Salary'].str.strip(' ').str.lstrip('$').str.rstrip('K').fillna(0).astype('int')
data['Max_Salary']=data['Max_Salary'].str.strip(' ').str.lstrip('$').str.rstrip('K').fillna(0).astype('int')
# int to string, fill zero ahead of 2 digits
cad['s']=cad['s'].astype(str).str.zfill(2)
# replace by the first column
df['First']=df['First'].str.split(expand=True)
# slice by the column df[].str.slice(start,stop,step)
df['sex'] = df['sexage'].str.slice(0,1)
df['age'] = df['sexage'].str.slide(1)

#extract digits from string
df['A'].str.extract('(\d+)')
df['A'].str.extract('(^\d+)')

# split digits in string
df['A'].str.split('(\d+)',expand=True)

# if contains string
df_fandry = df[df['log'].str.contains('3s',regex=True)]
# if not contains string
df_airdry = df[~df['log'].str.contains('3s',regex=True)]
df_airdry = df[~df['log'].str.contains('3s',regex=True,na=False)]

# change column datatype
df['slide'] = df['slide'].astype(int)

# replace
df = df.replace(np.nan, 'N/A', regex=True)
# choose random sample 
df.sample(n=100).describe() -> not allow duplicated row
df.sample(n=100,replace=True).describe() -> allow duplicated row
df.sample(frac=0.4)
df.sample(n=2,random_state=2)-> get the same row everytime
# stack -> transfer multiindex dataframe to flat 

# unstack -> transform a dataset with a pd.MultiIndex into a 2 dimensional array
map_2d=planck_map.unstack()
type(map_2d)

# multicolumn, muliindex flatten
dst = df[['1uL_wbc','sample','log']].groupby(['log','sample'],as_index=False).agg(['mean','std','count'])
dst['1uL_wbc','cv[%]'] = dst['1uL_wbc','std']/dst['1uL_wbc','mean']*100
# flat multi index
dst.reset_index(inplace=True)
# collape multi level columns 
dst.columns = [' '.join(col).strip() for col in dst.columns.values]


import matplotlib.pyplot as plt
# get corration
df_corr = df.corr()
f = plt.figure(figsize=(19,15))
plt.matshow(df.corr(),fignum=f.number)
plt.xticks(range(df.shape[1]), df.columns,fontsize=14,rotation=45)
plt.yticks(range(df.shape[1]), df.columns,fontsize=14)
cb = plt.colorbar()
cb.ax.tick_params(labelsize=14)
plt.title('Correlation Matrix', fontsize=16)
# alternative
styler=df.corr().style.background_gradient(cmap='coolwarm').set_precision(2)
styler=df.corr().style.background_gradient(cmap='coolwarm').set_properties(**{'font-size':'0pt'})
display(styler)


df.query()
df.query('country="China" and year == 2015')

#MultiIndex 
sample_multiindex = pd.MultiIndex.from_tuples([('A',1),('A',2),('A',3),('B',1),('B',1)], names=['letter','number'])
sample_series_with_multiindex=pd.Series(list(range(100,105)),index=sample_multiindex)

import pandas as pd
# dataframe from csv
# dataframe from dictionary
my_dict = {"tester": [1,2], "testers": [2,3]}
# key as column
df = pd.DataFrame.from_dict(my_dict)
# key as index
df = pd.DataFrame.from_dict(my_dict,orient='index')
df = pd.DataFrame.from_dict(my_dict,orient='index',columns=['a','b'])
df.to_csv("path and name of your csv.csv")
# set to data frame type to plot
df = pd.read_csv('ABC.csv') #read cvs
# reset column name at read_csv
df = pd.read_csv('ABC.csv',delimiter='\t',header=None,
        names=['n1','n3',*[f'g{i}' for i in range(19)]]) 

df2 = df.loc[0:12] #select rows
# get rid of "Unnamed:0"
# pd.to_csv(df,encoding='utf-8',index=True)
df = pd.read_csv('file.txt',index_col=[0])
# selective read
df = pd.read_csv('file.csv').query('year==1986')
# read first nrows 
df = pd.read_csv('file.csv',nrows=100)
# sort data
df.sort_values('age',inplace=True)
# set index as Date
df2['Date'] = pd.to_datetime(df2['Date'])
df2.set_index('Date', inplace=True) #set Date as index
df2['Close'].plot() #plot
# add 5 business days
import datetime 
from pandas.tseries.offsets import BDay
ts = pd.Timestamp(datetime.datetime.now())
>>> ts + BDay(5)

date_string = datetime.datetime.strptime("","")

hour_ago = (datetime.datetime.now() - datetime.timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")
now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

base = datetime.datetime.today()
date_list = [base - datetime.timedelta(days=x) for x in range(0, numdays)]

datetime.strptime : convert string to datetime
datetime.strftime : convert datetime to string

# conditionally multiply column
df.loc[df['vol']=='0.5uL',['rbc','wbc','plt']] *=2.0

# list operation: remove 
unwanted = [2, 3, 4]
item_list = [e for e in item_list if e not in unwanted] 

df.rename(columns={'close':ticker}, inplace=True)

df2.index=df2.index.values.astype('<M8[D]')

#
size=np.where(df['year']%10 ==0, 30,2)
color=np.where(df['country']=='China', 'red', 'orange')
ax = df.plot.scatter('x','y',logx=True,s=size,c=color)
df.plot.line(x='yy',y='life',ax=ax)

#list NaN
df_null = df[pd.isnull(df['info'])]
df.sample(10)
df.isnull().sum()

#sort index
df.sort_index(axis=0)
df.sort_index(axis=1)

# fillna with another column
df['info'].fillna(df['add_info'],inplace=True)

# find duplicated index
duplicated_idx =  df[df.index.duplicated()].index
df.drop(duplicated_idx, axis = 0, inplace=True)

df = df.loc[~df.index.duplicated(keep='first')]
# get unique label

#construct empty dataframe
df_empty = pd.DataFrame({'A': []})
return df_empty  #True
if df_empty.empty:
    pass

if df is None:
    pass


# tolist
slide_list = df['slide'].tolist()

#get row with index
df.log[idx]
#update colume with new values (no duplicated index)
df.update(slide_info)
df['sample'].update(slide_info['sample'])
# 
df['sample']=df['sample'].map(dict)

# merge two dataframe
c = pd.merge(df_left,df_right,on=key)
# concat two dataframe
c = pd.concat([df1,df2],axis=0,join='outer',sort=False)
# join two dataframe
df_ab = df_a.join(df_b, lsuffix='_A', rsuffix='_B')
df_ab = df_a.join(df_b, on='key', how='outer')
df_ab = df_a.join(df_b, on='key', how='inner')
# merge part of df_b to df_a
df_ab = df_a.merge(df_b['col1','col2'])

# transfer dataframe to dict
grps = df['group'].unique().tolist()
d_data = {grp:data['weight'][data['group']==grp] for grp in grps}
part_grp = data.groupby('group').size()[0]

# ANOVA with scipy
from scipy import stats
F,p = stats.f_oneway(d_data['ctrl'],d_data['trt1'],d_data['trt2'])

# ANOVA with scipy.stats
import scipy.stats as stats
stats.f_oneway(data_group1,data_group2,data_group3,data_groupN)
# ANOVA with statsmodels
import statsmodels.formula.api import ols
import statsmodels.api as sm
mod = ols('outcome_variable ~ C(groups_variable)',data=your_data).fit()
aov_table = sm.stats.anova_lm(mod, typ=2)
print(aov_table)
# ANOVA pairwise Comparisons
mod = ols('weight ~ group',data=data).fit()
pair_t = mod.t_test_pairwise('group')
pair_t.result_frame




#================================================== 
# numpy
#================================================== 
import numpy as np
# compute Covariance matrix
Sigma = np.cov(train_data, rowvar=1, bias=1)
# compute coordinate-wise variances in increasig order
coordinate_variances = np.sort(Sigma.diagonal())
# compute variances in eigenvector directions, in increasing order
eigenvector_vairances = np.sort(np.linalg.eigvalsh(Sigma))
#
eigenvalues, eigenvectors = np.linalg.eigh(Sigma)

np.random([0,1,4,5],size=n,p=[0.65,0.15,0.1,0.1]) # p proprotion to 
#list operation
x=arange(0,2*pi,2*pi/365)
v = []
v.append(np.array(cos(0*x))*c/sqrt(2))
v.append(np.array(sin(x))*c)

# list to array
A=vstack(v)
#
np.insert(arr,0,0)
np.argmin(arr,axis=0)
np.argsort()
np.argpartition()
np.stack(A)

np.nansum(mat)
np.nanmax(mat)
np.nanmin(mat)

ret = [2,3.4,54]
mean = np.mean(ret)


#transfer boolean into string
string_array = np.where(np.random.rand(5)>0.51,'Brown','Green')
# 
dict_vote = {'vote':np.where(np.random.rand(n) < 0.1,'Brown','Green')}
df_vote = pd.DataFrame({'vote':np.where(np.random.rand(n) < 0.1,'Brown','Green')})

#polynomial regression
z = np.polyfig(percX,percUp,degree)
f = np.poly1d(z)
x_new = np.linspace(0.1,2.1,50)
y_new = f(x_new)
plt.plot(x_new, y_new, 'r--')

#================================================== 
# jupyter
#================================================== 
# Markdown secetion
Double-click **here** for the solution.

<!-- The correct answer is:
It helps clarify the goal of the entity asking the question.
-->
# collapsible headings
$pip install jupyter_contrib_nbextensions
# write cell content to file
%%writefile ../lib/myfuncs.py
# load file to cell
%load ../lib/myfuncs.py

# add project folder to sys path
import sys
path ="" 
sys.path.append(path)
print(sys.path)

# Enable automatic reload of libraries
%load_ext autoreload 
%autoreload 2 # means that all modules are reloaded before every command
# load function from lib to execute but not display
from lib.funcs import *

# load py code 
%load filename.py
# write below line at the begining of the cell to write to py code
%%writefile filename.py
# run python code
%%run filename.py
# get help on magic function
%magic
%lsmagic
%magic_name
# run unix command
!python --version
!cat file.txt
# split cell
Ctrl+Shift+-
# merge cell
shift M
# inline images
%matplotlib inline


# ====================================
import plotly.express as px
import plotly.graph_objects as go

fig = go.Figure()
fig.add_trace(go.Bar(x=df.index, y=df['Min_Salary'],name='Min Salary'))
fig.add_trace(go.Bar(x=df.index, y=df['Max_Salary'],name='Max Salary'))
fig.update_layout(title='title',barmode='stack')
fig.show()
# ====================================

# 
0+0: reset kernel

# get help of modulur
# Shift+Tab

?function-name: shows the definition and docstring for that function
??function-name
doc(function-name)

%timeif

%debut

img_path = Path('')
path.ls()

#make dir in current working_folder
path = Path('new/folder_1')
folder = 'black'
dest = path/folder
dest.mkdir(parents=True,exist_ok=True)


# hide warning
import warnings
warnings.filterwarnings('ignore')
#================================================== 
from tqdm import tqdm
#================================================== 
#Os
import os
ticker_list = os.listdir(statspath)
os.remove(f"{statspath}/.DS_Store")
ticker_list.remove(".DS_Store")

os.makedirs(stock_dfs)
os.path.isfile(filename)
os.path.exists(filename)
os.path.join(path,"zlog")+"\\"
os.listdir(path)
os.remove('file.txt')
os.path.isdir(folder)
#get path pwd
path=os.path.abspath(os.getcwd())
#get path python
os.path.realpath(sys.executable)
# get paht 1 level up
path = os.path.abspath(os.path.join(os.getcwd(),".."))
# get paht 2 level up
path = os.path.abspath(os.path.join(os.getcwd(),"../.."))

# split path
nb_dir = os.path.split(os.getcwd())

# rename folder
os.rename(os.path.join(base_folder,old),os.path.join(base_folder,new))

# rename a file
os.rename(src,dst)

# copy files 
import shutil
shutil.copy2(f_from,f_to)

my_png = os.path.join(path,"my_png")+"\\"
os.listdir(my_png)
#alternative
file_list = glob.glob(my_png+"_10x_performance.txt")

#
myPath = Path('F:\mydrive\as\fjk.jpg')
myPath.ls()
myPath.parent.stem ->as
myPath_sub1 = myPath/'sub1'

# =======================================================
# List comprehension
# =======================================================
# find duplicated elements in a list
l = [1,1,2,2,3,3,4,5]
set([x for x in l if l.count(x)>1])

#reverse a list
list1 = [i for i in range(10)]
list1_rev = list1[::-1]

#list only files
file_lsit = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder,f))]

a = ['a','b','c']
maxlen = 10
result = [(0 if i+1 > len(a) else a[i]) for i in range(maxlen)]

# find common elements in two lists
common = [i for i in x if i in y]

# replace none alnum with '_'
ff = [c if c.isalnum() else '_' for i in title]

# remove the last element in the list
my_list.pop()
# remove the first element in the list
my_list.popleft()
# remove element from list
list.remove(element)

#
# list comprehension
# transfer all string in a list to int
int_list = list(map(int, string_list))

# check empty list
a = []
if not a:
    print(f'empty')


#===============================
# convert list to dictionary keys
list_samples
list_donors
dict_sample = dict.fromkeys(list_samples,1.0)
# create dict from 2 lists
zipObj= zip(list_samples,list_donors)
dict_sample = dict(zipObj)
df['donor']=df['sample'].map(dict_sample)

# melt columns to rows
df_melt = pd.melt(df,id_vars='grp_pair',value_vars=cols,value_name='p_value')
#concat two dictionary
d1.update(d2)
#or
dboth={}
dboth.update(d1)
dboth.update(d2)

# print dictionary
for key,item in res_dict.items():
    print(key,item)

# create list of nan
nan_list = [np.nan for i in range(9)]

#list
a = []
a.append(b)
sorted(a)
sorted(a,reverse=True)
#map list using dictionary
l_map = list(map(my_dict.get,low_mag))

# combine two lists to one
newlist = list1 + list2
nwelist = [*list1, *list2]


#================================
# plot
#================================
colormap = sns.diverging_palette(220,10,as_cmap=True)
colormap=plt.cm.RdBu 

current_palette = sns.color_palette()
sns.scatterplot(x='x',y='y',data=df,color=current_palette[0])

# save figure to pdf
%pylab inline
from matplotlib.backends.backend_pdf import PdfPages
from lib.PlotTime import PlotTime
pp = PdfPages('MemoryBlockFigure.pdf')

figure(figuresize=(6,4))
Colors='bgrcmyk'    # color
lineStype=['-']
for m_i in range(len(m_list)):
    Color = Colors[m_i % len(m_list)]
    PlotTime()

plt.savefig('fivure.png',bbox_inches="tight")
if not show:
    plt.close()


# x label overlap
fig, axs = plt.subplots(figsize=(6,4))
axs.hist(x)
axs.set_xticks(axs.get_xticks()[::2])
plt.show()

# scientific label format
axs.ticklabel_format(axis='x',style='sci',scilimits=(0,2))

# horizontal line
plt.axhline(y=1,color='gray',linestyle='--')
ax.axhline(y=1,color='red',linestyle='-')

#plot settings
# square plot
ax.axis('equal')
ax.get_legend().set_visible(False)
ax.add_legend(title='title')

# remove legend title
h,l = ax.get_legend_handles_labels()
ax.legend(h[1:],l[1:],ncol=3,loc='upper center',
        bbox_to_anchor=(0.5,1.25),
        columnspacing = 1.3,labelspacing = 0.0,
        handletextpad = 0, handlelength = 1.5,
        fancybox=True, shadow=True)
        


# 
plt.rcParams.update({'figure.figsize':(8,6),'figure.dpi':120})
plt.figure(figsize=(4,4))

# pillow image compression
from PIL import Image
img = "myimg.png"
im = Image.open(img)
im_resize = im.resize((600,600),Image.ANTIALIAS)    # compression optimized
im_resize.save("resized_"+img)
im_resize.show()    # external window
display(im_resize)  # jupyter

#conver between a PIL image and a numpy array
em = Image.open("sample.png")
np_im = numpy.array(im)

new_im = Image.framarray(np_im)

# create new empty image
from PIL import Image
image = Image.new('RGB',(800,1280),(255,255,255))
img.save('image.png','PNG')


#===============================
# get rid of \n
.rstrip()
# String.strip() -> remove leading and trailing white spaces
.strip()



#===============================
# arg
args = [i for i in my_list]
F,p = stats.f_oneway(*arg*args)
#===============================
# format
form = '.2f'
splot.annotate(format(p.get_height(),form))
# left align
print('{:20s}'.format('Align Left'))
print('{:>20s}'.format('Align Right'))
print('{:^20s}'.format('Align center'))

# print list with format
grp_format = " ".join(["{:>10s}".format(x) for x in x_list])
print('{}'.format(grp_format))

# string format
print('{:5d}'.format(1))-> 00001

print('stat=%.3f, p=%.3f'%(stat,p))

print('Hello, %s' %str)

#===============================
# pandas multiindex
# resulting a multi-index dataframe
dst = df[['log','sample','rbc']].groupby(['log','sample'],as_index=False).agg(['mean','std'])
dst.columns
# get one column
dst['rbc','mean']
# add new column
dst[('rbc','cv')]=dst['rbc','std']/dst['rbc','mean']*100

#===============================
# drop non numeric rows
ds = ds[pd.to_numeric(ds['slide'],errors='coerce').notnull()]

#===============================
#Jupyter notebook
#list magic command
%lsmagic
%debug
%timeit
#-------------------------------
%%latex
$e^{i pi} = -1$
#-------------------------------
# get build-in __doc__
# add ? after fuction name
FuncName?
FundName.__doc__
#-------------------------------
# set envirnment variables
%env OMP_NUM_THREADS%env OMP_NUM_TREADS=4
#-------------------------------
#timing
import time
#return information about a single run of the code in one cell
%%time
#-------------------------------
# comment multiple lines
Ctrl+/



#uses the Python timeit module which runs a statement 100,000 time
#(by default) and provides the man of the fastest three times.
%timeit
#-------------------------------
#shows (in a popup) the syntax highlighted contents of an external fiel
%pycat xx.py
#-------------------------------
#debug tool
%pdb
#-------------------------------
#suppress output by adding ;
#-------------------------------
#execute shell commands
!ls *.csv
!pip install numpy
!pip list |grep pandas
#-------------------------------
hold <Alt> column selection


#===============================
# python 2
# virtual environment
$pip install virtualenv
# add virtualenv pytest_27_venv
$virtualen pytest_27_venv
# activate pytest_27_venv
$source ./pytest_27_venv/bin/activate
#install packet
$pip install pytest
$deactivate
# deleate the top level directiory 

# python 3 build-in virtual envio
# create virtual environment
$python3 -m venv pytest_3_vene
$source pytest_3_venv/bin/activate
$pip install pytest
$deactivate


if __name__ =="__main__":
    print("Hello World!")

#===============================
# df to excel
#===============================

from openpyxl.workbook import Workbook
to_excel = df.to_excel('df.xlsx')

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
wb = Workbook()
ws = wb.active

ws2 = wb.create_sheet('new sheet')

ws.title='MySheet'
rows = dataframe_to_rows(df,index=False)
for i, row in enumerate(rows,1):
    for j,col in enumerate(row,1):
        ws.cell(row=i,column=j,value=col)
wb.save('file.xlsx')




#===============================

#===============================
# print dict
data={'city':1,'region':12,'country':23}
print("{city}, {region},{country}".format(**data))

#===============================
$pip install requests
$pip install numpy simpleaudio
# opencv
$pip install opencv-contrib-python imageio
# PEP python enhance protocol 
# pycodestyle: check python code against style convention
$pip install pycodestyle
$pycodestyle code.py
# flake8: comtinges a debutter, pyflates with pycodestyle
$pip install flake8
$flake8 code.py
# auto formatters
$pip install black
$black code.py
# alter linelength limit
$black --line-length=79 code.py

#===============================
file_name[-3:]=='jpg'
if file_name.endswith('jpg'):
    print(file_name)
file_name[:3]=='jpg'
if file_name.startswith('jpg'):


#===============================
#unit test pandas
import pandas as pd
pd.testing.assert_frame_equal(my_df,expected_df)
pd.testing.assert_series_equal(my_series,expected_series)
pd.testing.assert_index_euqal(my_index,expected_index)

#===============================
# python program strucutre
# smart_door.py
def close():
    pass
def open():
    pass

import smart_door
smart_door.open()
smart_door.close()
# alternative
from smart_door import open
open()

#===============================
#bootstrap: sample iwth replacement
df.sample(100)-> randomyl choose 100 instances from df
df.sample(100,replace=True).describe()
bootstrap=pd.DataFrame({'meangrad':[df.sample(100,replace=True).grade.mean() for i in range(1000)]})

# get pd quantile
df['meangrade'].quantile(0.025)


#===============================
import scipy
#normal distribution
n = scipy.stats.norm(mu,sigma)
scipy.stats.percentileofscore(df['count'],340) -> return confidence p of getting 340
100-scipy.stats.percentileofscore(df['Brown'],0.511)

from scipy.stats import itemfreq
fg = itemfreq(list(all_chat_list)[0].split(' '))
fg = fg[fg[:,1].astype(float).argsort()][::-1]
print(fg[1:4])


# ANOVA
from scipy import stats
F,p = stats.f_oneway(d_data['ctrl'],d_data['trt1'],d_data['trt2'])

#===============================
# plot using seaborn
import seaborn as sns
import matplotlib.pyplot as plt

# 'paper','talk','notebook'
sns.set_context('paper')
sns.set_color_codes('pastel')
sns.set_color_codes('muted')
sns.set(style="darkgrid")

sns.set(style='whitegrid', color_codes=True, font_scale=1.7)

sns.scatterplot(x="total_bill",y="tip",hue="size",style="smoke",data=df)

sns.catplot(x='speed',y='tc',hue='sample',col='type',col_wrap=2,data=df)
g.set_axis_labels("","Totla number of TargetCells")
plt.subplots_adjust(top=0.8)
g.fig.suptitle('Title')
g.set_ylabels('ylabel')

g = sns.catplot(x='class',hue='who',col='survived',data=titanic,kind='count',height=4,aspect=.7)
(g.set_axis_labels('xlabel','ylabel')
        .set_xticklabels(['','',''],rotation=90)
        .set_title('{col_name}{col_var}')
        .set(ylim=(0,1))
        .despine(left=True))
sns.set_context('paper',fontsize=1.2)

g.set_xticklabels(g.get_xticklabels(),rotation=90)

# lmplot
# use 'hue' argument to provide a factor variable
sns.lmplot(x='x',y='y',data=df,fit_reg=False,hue='species',legend=False,palette='Set2')

g = sns.lmplot(x='total_bill',y='tip',row='sex',col='time',
        data=tip, height=3)
g = (g.set_axis_labels('Total bill','Tip')
        .set(xlim=(0,60),ylim=(0,12),
            xticks=[10,30,50],yticks=[2,6,10])
        .fig.subplots_adjust(wspace=.02)
        )

# swarmplot
g = sns.swarmplot(x='category',y='sodium',data=df)
# histogram
from scipy.stats import norm
sns.distplot(df['x'],fit=norm, kde=False,vertical=True)

df_grouped = df.groupby(by='DEATH_EVENT')
sns.distplot(df_grouped.get_group(0)[title],bins=10,ax=ax,label='No')
sns.distplot(df_grouped.get_group(1)[title],bins=10,ax=ax,label='Yes')

# joint histogram
g = sns.jointplot(x='x',y='tip',data=df)
g.fig.suptitle('title')
g.fig.tight_layout()
g.fig.subplots_adjust(top=0.95)

g = sns.jointplot(x='x',y='y',data=df,kind='kde',space=0,color='g')
g=(sns.jointplot(x='hct',y='wbc',data=df,color='g').plot_joint(sns.kdeplot, zorder=0,n_levels=20))

# kde plot
sns.kdeplot(x,y,cmap=cmap,cut=5,shade=shade)
sns.kdeplot(x,y,cmap=cmap,cut=5,shade=shade,ax=ax)

# regression plot
# ci: 68% confidence interval
sns.regplot(x='x',y='y',data=df,color='g',marker='+',ci=68,order=2)
        

#move legend outside
plt.legend(title='mytitle',bbox_to_anchor=(1.05,1),loc=2)
plt.legend(bbox_to_anchor=(1.05,1),loc=2,borderaxesspad=0.)
plt.legend(bbox_to_anchor=(1.05,1),ncol=1,loc=2,borderaxesspad=0.)

# use unique legend
handles, labels = plt.gca().get_legend_handles_labels()
by_label = dict(zip(labels,handles))
plt.legend(by_label.values(),by_label.keys(),bbox_to_anchor=(1.05,2),loc=2)

by_label={}
handles, labels = axs[i].get_legend_handles_labels()
by_label.update(dict(zip(labels,handles)))


plt.scatter(df_red['x'],df_red['y'],s=df_red['count'])

# add margin to plot
plt.margins(1)

#plot 
plt.plot(X,Y,'.')

sns.jointplot(x="x",y="y",data=df)
sns.jointplot(x="x",y="y",data=df,kind="kde")

#QQ plot
#https://seaborn-qqplot.readthedocs.io/en/latest/
import matplotlib.pyplot as plt
import seaborn as sns
import seaborn_qqplot as sqp
iris = sns.load_dataset('iris')
sqp.qqplot(iris,x="sepal_length",y="petal_length",height=4,aspec=1.5,)

## Pairplot
iris = sns.load_dataset('iris')
g = sns.pairplot(iris,hue='species',palette='husl',markers='d',size=2.5,
        plot_kws={'s':40,'alpha':1.0,'lw':0.5,'edgecolor':'k'},vars=['length','width'])
handles = g._legend_data.value()
labels = g._legend_data.keys()
g.fig.legend(handles=handles,labels=labels,loc='upper center', ncol=1)
g.fig.legend(handles=handles,labels=labels,loc='lower center', ncol=3)
g.fig.legend(handles=handles,labels=labels,loc='upper left', ncol=3)
g.fig.subplots_adjust(top=0.92,bottom=0.08)
g.fig.suptitle('title',y=1.08)

# remove unneeded (last) subplots
axs.flat[-1].remove()
axs.flat[-2].remove()

## Joinplot -> shows coorelation with p-value
g = sns.joinplot(x='Protein',y='Total Fat',data=df)

#
matplotlib.rc('xtick',labelsize=14)
matplotlib.rc('ytick',labelsize=14)

# plot histgram
# only col
grid = sns.FacetGrid(train_df, col='Survived')
grid.map(plt.hist,'Age',bins=20, density=1, alpha=0.5)
# col and row
grid = sns.FacetGrid(train_df, col='Survived', row='Pclass', size=2.2, aspect=1.6)
grid.map(plt.hist, 'Age', alpha=0.5, bins=10)
#
d = {'color':['red','blue'],'ls':['--','-'],'marker':['v','^']}
g = sns.FacetGrid(df,col='sample',hue='sample',hue_kws=d,col_wrap=2)
g.map(sns.scatterplot,'under_stain','over_stain')
#
# using pointplot
# palette = 'deep','muted','pastel','bright','dark','Set1','Set2','husl',
# 'hsv','rainbow'
grid = sns.FacetGrid(train_df, col='Embarked', size=2.2, aspect=1.6)
grid.map(sns.pointplot, 'Pclass', 'Survived', 'Sex', palette='deep', order=[1,2,3], hue_order=['female','male'])
grid.add_legend()
#
# lineplot
sns.lineplot()
# barplot
sns.barplot(train_df['Sex'],train_df['Survived'])
#ci= None, or sd 
grid = sns.FacetGrid(train_df, row='Embarked', col='Survived', size=2.5, asepect=1.6,)
grid.map(sns.barplot, 'Sex','Fare', order=['female','male'],alpha=.5, ci=None)

sns.barplot(x='Accuracy mean', y='MlA', data=MLA_compare, coloer = 'm')
sns.barplot(x='Accuracy mean', y='MlA', data=MLA_compare, coloer = 'm',ci=None)
sns.barplot(x='Accuracy mean', y='MlA', data=MLA_compare, coloer = 'm',ci=None,label='',edgecolor='')
# countplot
sns.countplot(x='sample',data=df)
# Title
plt.title('Machinea \n')
plt.xlabel('Accuracy Score ')
plt.ylabel('Algorithm')

# sns annotation
splot = sns.boxplot(x='x',y='y',data=df)
form = '2d'
form = '.2f'
for p iin splot.patches:
    splot.annotate(format(p.get_height(),form),
            (p.get_x()+0.5*p.get_width(),p.get_height()+1),
            ha='center',
            va='center',
            xytext=(0,10),
            textcoords='offset points',
            size=8)

#================================================== 
# matplotlib

# subplots annotate
labels = df['label']
fig,ax = plt.subplots()
for i in label in enumerate(labels):
    ax.annotate(label, (X.iloc[i],y.iloc[i]))

# another subplots
plt.figure(1,figsize=(15,6))
n=0
for x in ['age','income','spend']:
    n += 1
    plt.subplot(1,3,n)
    plt.subplots_adjust(hspace=0.5, wspace=0.5)
    sns.distplot(df[x],bins=20)
    plt.title('Distplot of {}'.format(x))
plt.show()


# flexible use of subplot
# option 1
fig = plt.figure(figsize=(10,14))
ax1 = fig.add_subplot(211)   # (row, col, current_plot_location)
ax1.plot(x,y,'k-',label='jkld')
ax1.set_ylim((0,0.2))
ax1.set_xlable('n_estimater')
ax1.set_ylable('error rate')
leg = ax1.legend(loc='upper right',fancybox=True)
leg.get_frame().set_alpha(0.7)
ax2 = fig.add_subplot(212)   # (row, col, current_plot_location)
ax2.plot(x,y,'k-',label='jkld')
ax2.set_ylim((0,0.2))
ax2.set_xlable('n_estimater')
ax2.set_ylable('error rate')
leg = ax1.legend(loc='upper right',fancybox=True)
leg.get_frame().set_alpha(0.7)
plt.show()

# option 2
plt.figure(figsize=(14,12))
plt.subplot(2,1,1)
plt.plot(x1,y1,'o-',label='test data')
plt.title('title 1')
plt.xlabel('xlabel')
plt.ylabel('ylabel')
plt.legend()
plt.grid()
plt.grid(axis='x')
plt.subplot(2,1,2)
plt.plot(x2,y2,'o-',label='training data')
plt.title('title 2')
plt.xlabel('label 2')
plt.show()
plot_color = 'br'
class_names='AB'
for i, n, c in zip(range(2), class_names, plot_colors):
    idx = np.where(y==1)
    plt.scatter(X[idx,0],X[idx,1],
                c = c, cmap = plt.cm.Paired,
                s=20, edgecolor='k',
                label='Class %s' % n)
# option 3
fig, saxis = plt.subplots(2,2,figsize=(16,12))
sns.barplot(x='Embarked',y='Survived',data=data1, ax=saxis[0,0])
axs[0,0].set_xlabel('x')
axs[0,0].set_ylabel('x')
sns.barplot(x='Pclass',y='Survived',data=data1, ax=saxis[0,1])
sns.barplot(x='Embarked',y='Survived',data=data1, ax=saxis[1,0])
sns.barplot(x='Pclass',y='Survived',data=data1, ax=saxis[1,1])
# option 4
fig, (ax1,ax2) = plt.subplots(2,1,figsize=(14,10))
sns.boxplot(x='Pclass',y='Fare', hue='Survived',data=data1, ax=ax1)
ax1.set_title('title 1')
sns.violinplot(x='Pclass',y='Age',hue='Survived',data=data1, ax=ax2)
ax2.set_title('title 2')
# option 5
fig, ax = plt.subplots(2,1,figsize=(14,10))
sns.boxplot(x='Pclass',y='Fare', hue='Survived',data=data1, ax=ax[0])
ax1.set_title('title 1')
sns.violinplot(x='Pclass',y='Age',hue='Survived',data=data1, ax=ax[1])
ax2.set_title('title 2')

# hide legend
axs.get_legend().set_visible(False)
# remove legend
axs.legend_.remove()

# equal
axs[i,j].settitle('title',fontsize=10)
axs[i,j].axis('equal')
axs[i,j].set_aspect('equal','box')

# plot meshgrid
x_min, x_max = X[:,0].min()-1, X[:,0].max()+1
y_min, y_max = X[:,1].min()-1, X[:,1].max()+1
xx,yy = np.meshgrid(np.arange(x_min, x_max, plot_step),
                    np.arange(y_min, y_max, plot_step))
Z = dbt.predict(np.c_[xx.ravel(),yy.ravel()])
Z = Z.reshape(xx.shape)
cs = plt.contourf(xx,yy,z,cmap=plt.cm.Paired)

# plot diagonal line
ax.plot(ax.get_xlim(),ax.get_ylim(),ls='--',c='red')

# subplot title
fig,axs = plt.subplots(2,1,sharex=True)
st=fig.suptitle('title',fontsize='x-large')
st.set_y(0.9)
st.set_x(0.45)
axs[0].set_title('title1')
axs[1].set_title('title1')

fig,axs = plt.subplots(2,1,sharex='col',sharey='row')

# rotate xticks
g = sns.boxplot(x='log',y='count',data=df)
g.set_xticklabels(g.get_xticklabels(),rotation=90)
ax.set_xticklabels(ax.get_xticklabels(),rotation=90)

# matplotlit
plt.xticks(rotation=90)
plt.xticks(df.index[::10],rotation=90)

# invert plot axis
ax.invert_xaxis()
ax.invert_yaxis()

#  adjust ticks xtick, ytick
from matplotlib import ticker
tick_locator = ticker.MaxNLocator(10) #max number of tickers
tick_locator = ticker.MultipleLocator(1) # 

ax.xaxis.set_major_locator(tick_locator)

#================================================== 
# undo seaborn set_style
import matplotlib as mpl
mpl.rcParams.update(mpl.rcParamsDefault)
#assign same color palette to different plot
sns.set_style('white',font_scale=1.2)
pal = sns.color_palette('Paired')

sns.boxplot(x='site',y='value',hue='label',data=df,
        palette=pal,fliersize=0)
sns.stripplot(x='site',y='value',hue='label',data=df,
        jitter=True,split=True,linewidth=0.5,palette=pal)
plt.legend(loc='upper left')

#specify standard color to label 
pal = {'healthy':'green','disease':'blue'}

#specify specified color to label
dark_brown = '#B25116'
dark_pink = '#FB84D1'
pal = {'healthy': dark_brown,'disease': dark_pink}

# flat axes in subplot
f, axs = plt.subplots(3,3,figsize=(9,9),sharex=True,sharey=True)
for ax in axs.flat:
    sns.scatterplot(x,y,ax=ax,data=df)

# use log scale
splot = sns.scatterplot(x='x',y='y',data=df)
splot.set(xscale='log')
splot.set(yscale='log')

# pass setting using dict
fig,ax = plt.subplots()
stripplot_kwargs = dict({'linewidth':0.6, 'size':6,'alpha': 0.7}, **kwargs)
sns.stripplot(x='x',y='value',hue='label',data=df,fliersize=0,ax=ax,**kwargs)
ax.legend_.remove()

# multiple plots on same chart
fig, ax = plt.subplots()
ax.plot(df_real['pred'],label='predict',color='red')
ax.plot(df_real['price'],label='real',color='darkgreen')
ax.set_xticks(ax.get_xticks()[::5])

# legend -> keep first 2 legend
handles, labels = ax.get_legendhandles_labels()
lgd = ax.legend(handles[0:2], labels[0:2],
        loc = 'upper left',
        fontsize = 'large',
        handletextpad = 0.5)
lgd.legendHandles[0]._sizes=[40]
lgd.legendHandles[1]._sizes=[40]

# add text to plot
sns.scatter(x='time',y='printtime',data=di,ax=ax)
for idx in di.index:
    ax.text(di.at[idx,'time'],di.at[idx,'printtime'],idx)


#================================================== 
# plt vertical line
# c for color
plt.axvline(df['income'].mean(),c='C1')
plt.axvline(df['income'].median(),c='C1',linestyle='--')
plt.axvline(df['income'].quantile(0.24),c='C1',linestyle=':')
plt.axvline(df['income'].quantile(0.75),c='C1',linestyle=':')

# vline
plt.vlines(x,ymin,ymax,colos,linestyles)


#================================================== 
#pandas plot
df['income'].plot(kind='hist',histtype='step',bins=3,density=True)
df['income'].plot.density(bw_method=0.5)
plt.axis(xmin=0,xmax=3)

df['income'].plot(kind='hist',histtype='step',bins=20,density=True)
levels=[0.25,0,1,2,3]
plt.xticks(np.logs10(levels),levels) #(location,label)




#===============================
# dataframe plot
#simplest matplot example
import matplotlib.pyplot as plt
df.plot()
df.plot(label='X={} deg'.format(X))
plt.title('Emission at different latitudes')
plt.legend();

#===============================
#read png
img = matplotlib.image.imread('london.png')
plt.imshow(img,extend=[-0.38,0.38,-0.38,0.38]) # resize
plt.scatter(df['x'],df['y'],color='b')
plt.imshow(img,origin='upper',interpolation='none',
        cmap=plt.get_cmap('gray'),
        extend=[135.3,34.2,344,33])
#===============================

#===============================
# fastai
from fastai.vision import *
#===============================
path_img = Path(':\F:')
learn.export(file='export_zc.pkl')
learn = load_learner(path_img, file='export_zc.pkl')

path_img = Path(':\F:')
img_list = ImageList.from_folder(path_test)
img_list.items # file name

learn = load_learner(path_model, test=ImageList.from_folder(path_test))
preds, y = learn.get_preds(ds_type=DatasetType.Test)


fnames = get_image_files(path_img)

img = open_image(path/'black'/'jio.jpg')

#  Workflow
#
folder = 'black'
dest = path/folder
dest.mkdir(parents=True,exist_ok=True)

classes = ['teddy','grizzly','black']
#download images
download_images(path/file, dest, max_pics = 200)
download_images(path/file, dest, max_pics=20, max_workers=0)
# verify images
for c in classes:
    verify_images(path/c, delete=True, max_workers=8)

np.random.seed(42)
data = ImageDataBunch.from_folder(path, train=".",  # training is the same folder
        valid_pct = 0.2,                        # 20% validation set
        ds_tfms=get_transforms(),
        size=224,
        num_workers=4).normalizie(imagenet_stats)
'''
data = ImageDataBunch.from_csv(path, folder='.', valid_pct=0.2, csv_labels='cleaned.csv',
        ds_tfms=get_transforms(),size=224,num_workers=4).normalize(imagenet_stats)
'''
data.classes
data.show_batch(row=3,figsize=(7,8))
data.show_batch(2,figsize=(10,7), ds_type=DatasetType.Valid)
data.classes, data.c, len(data.train_ds), len(data.valid_ds)
#learn = create_cnn(data, models.resnet34, metrics=error_rate)   # create_cnn is deprecated
learn = cnn_learner(data, models.resnet34, metrics=error_rate)   #get pre-trained resnet34
learn = cnn_learner(data, models.resnet34, metrics=error_rate,pretrained=False) #get rid of pre-trained resnet34
learn.fit_one_cycle(4)
learn.save('stage-1')   # workflow 1
#plot losses
learn.recorder.plot_losses()
# re-train start

# evaluate results
interp = ClassificationInterpretation.from_learner(learn)
interp.plot_top_losses(9, figsize=(15,11))
interp.plot_confustion_matrix()
interp.most_confused(min_val=2)

# Clean up, delete confusing images
from fastai.widgets import *
losses,idxs = interp.top_losses()
#note all data has in fastai has .x and .y class
# data.valid_ds, data.train_ds, data.test_ds
top_loss_paths = data.valid_ds.x[idxs]  # x : image path and name , y: image class/label
fd = FileDeleter(file_paths=top_lass_paths)

# Clean data
from fastai.widgets import *
# read all data
data = (ImageList.from_folder(path)
        .split_none()
        .label_from_folder()
        .transform(get_transforms(),size=224)
        .databunch())
learn_cln = cnn_learner(db, models.resnet34, metrics=error_rate)
learn_cln.load('stage-2')
ds,idxs = DatasetFormatter().from_toplosses(learn_cln)
ds,idxs = DatasetFromatter().from_similars(learn_cln)
ImageCleaner(ds,idxs,path,duplicates=True)


# ipywidget 

# production
learn.export()
learn = load_learner(path)
img = open_image(path/'x.jpg')
pred_class, pred_idx, outputs = learn.predict(img)
pred_class

# 
classes = ['black','grizzly','teddy']
data2 = ImageDataBunch.single_from_classes(path, classes, tfms=get_transforms(),size=224).normalize(imagenet_stat)

# garbage collect
gc.collect()

# tips
train_loss < valid_loss => increase learning rate or more epoch
# too many epochs -> over fitting

#split option
https://docs.fast.ai/data_block.html
split_by_rand_pct(valid_pct=0.1,seed=42)
split_by_folder
split_by_idx(valid_idx=range(800,1000))

# ===========================================
# create data bunch
data = (ImageFileList.from_folder(path) #Where to find the data? -> in path and its subfolders
        .label_from_folder()             #How to label? -> depending on the folder of the filenames
        .split_by_folder()               #How to split in train/valid? -> use the folders
        .add_test_folder()               #Optionally add a test set
        .datasets()                      #How to convert to datasets?
        .transform(tfms, size=224)       #Data augmentation? -> use tfms with a size of 224
        .databunch(bs=8))                #Finally? -> use the defaults for conversion to ImageDataBunch
        )
data.train_ds[0]
data.show_batch(row=3, figsize(5,5))

# Regression model
# TargetCell    
data = (ImageList.from_folder(path_img) #Where to find the data? -> in path and its subfolders
        .split_by_rand_pct(0.15)
        .label_from_func(get_float_labels, label_cls=FloatList)
        .transform(tfmss, size=[448,608])
        .databunch(bs=bs)).normalize(imagenet_stats)
# re-train
learn = cnn_learner(data, models.resnet18, metrics=rmse)    # error metrics tell regression
learn = cnn_learner(data, models.resnet18, metrics=rmse).load('stage_1')

learn.show_results()


# ===========================================
def acc_02(inp,targ): return accuracy_thresh(inp,targ,thresh=0.2)
acc_02 = partial(accuracy_thresh, thresh=0.2)    # call accuracy_thresh with argument thresh = 0.2

learn.fit_one_cycle(5,slice(0.01))


# ===========================================
# split the data source to several data bunch
np.random.seed(42)
src = (ImageFileList.from_folder(path),
        .label_from_csv('train_v2.csv',sep=' ', folder='train-jpg', suffix='.jpg')
        .random_split_by_pct(0.2))

size = src_size//2      # // : integer divide 
# reduce size of data
data = (src.transform(tfms, size=128)      # image size [128x128]
        .databunch(bs=bs).normalize(imagenet_stats)
# train the learner
learn.save('stage-1')
# replace the learner data with new data bunch
data = (src.dataset(),
        .transform(tfms, size=256, tfm_y=True) # flip over y-axis, tfm_x=True
        .databunch(bs=bs).normalize(imagenet_stats)
learn.data = data       # replace the data in learner
learn.freeze()          # only train the last few layers, keep the majority of info from previosu learner
learn.lr_find()
learn.recorder.plot()
lr = lr/2
learn.fit_one_cycle(5,slice(lr))        # train last few layers
learn.git('stage-2')

learn.unfreeze()        # unfreeze all
learn.fit_one_cycle(5, slice(1e-5, lr/5))

learn.freeze_to(-2)      # unfreeze last 2 layers
learn.fit_one_cycle(1,slice(1e-3),moms=(0.8,0.7))   # moms : decrease momentum
learn.save('second')
learn.load('third')
learn.freeze_to(-3)      # unfreeze last 2 layers
learn.fit_one_cycle(1,slice(1e-3),moms=(0.8,0.7))

# ===========================================
data.c
len(data.train_ds)
len(data.valid_ds)
#replace learn data
learn.data = data
learn.fit_one_cycle()

#plot loss
learn.recorder.plot_losses()


# code snippet
# ===========================================
test = ImageList.from_folder(path/'test-jpg').add(ImageList.from_folder(path/'test-jpg-additional'))
learn = load_learner(path_model, file=file_model, test=test_imglist)
preds, _ = learn.get_preds(ds_type=DatasetType.Test)
thresh = 0.2
labelled_preds = [' '.join([learn.data.classes[i] for i,p in enumerate(pred) if p > thresh]) for pred in preds]
fnames = [f.name[:-4] for f in learn.data.test_ds.items]
df = pd.DataFrame({'image_name':fnames, 'tags':labelled_preds}, columns = ['image_name','tags'])
df.to_css(path/'submission.csv',index=False)
# ===========================================
# dummy transfer boolean to number
df2= pd.get_dummies(df,columns=['MBA'], drop_first=True)
# ===========================================
get_y_fn = lambda x: path_lbl/f'{x.stem}_P{x_suffix}'


open_image(img)
open_mask(img)
# ===========================================
free = gpu_mem_get_free_no_cache()
if free > 8200: bs=8
else:           bs=4
print(f'using bs={bs}, have {free}MB of GPU RAM free')
# ===========================================
#segregation model
learn = Learner.create_unet(data, models.resnet34,metrics=metrics)
learn = Learner.create_unet(data, models.resnet34,metrics=metrics).to_fp16()    # 16bit precision?
# ===========================================
lr=3e-3         #rule of thumb

# python
size = src_size//2  # integre divide

# ===========================================
learn.show_results()

# muliti-label classification accuracy 
accuracy_thresh(inp, targ, thresh=0.2)

# ===========================================
# inspect images
fnames = get_image_files(path_img)
fnames[:1]
img_f = fnames[0]
img = open_image(imf_f)
img.show(figsize=(5,5))


# ===========================================
learn.save('stage-1')
learn.load('stage-1')
learn.show_results()

# ===========================================
# define mean square error
learn = create_cnn(data, models.resnet34)
learn.loss_func = MSELossFlat() #mean square error

# ===========================================
data = (TextList.from_csv(path, 'texts.csv',cols='text')
        .split_from_df(cols=2)
        .label_from_df(cols=0)
        .databunch())

data_lm = (TextList.from_folder(path)
           #Inputs: all the text files in path
            .filter_by_folder(include=['train', 'test']) 
           #We may have other temp folders that contain text files so we only keep what's in train and test
            .random_split_by_pct(0.1)
           #We randomly split and keep 10% (10,000 reviews) for validation
            .label_for_lm()           
           #We want to do a language model so we label accordingly
            .databunch(bs=bs))
data_lm.save('tmp_lm.pkl')
# ===========================================
# data bunch operation
data.save()
data.load()
data.show_batch()

data.train_ds[0][0]
data.valid_da[][]
# ===========================================
open('file.txt','r').readlines()[:10]
# ===========================================
df = pd.readcsv(path/'adult.csv')
test = TabularList.from_df()
cat_names = ['workclass','education','material_status']  #categorical
cont_names = ['age','fnlwgt']   # continuous data
pros = [FillMissing, Categorify, Normalize] #pre-processor to prepare data
test = TabularList.from_df(df.iloc[800,1000].copy(),path=path,cat_names=cat_names, cont_names=cont_names)
data = (TabularList.from_df(df, path=path, cat_names=cat_names,cont_names=cont_names,procs=procs)
        .split_by_idx(list(range(800,1000)))
        .add_test(test, label=0)
        .databunch())

# ===========================================
# Tabular_learner
learn = tabular_learner(data, 
                       layers=[300,100], #two cycles with 300 rows then 100 rows in the matrix 
                       #emb_szs = xxx, #embedding defaults usd
                       metrics=rmse, #metric of interest
                       ps = [.1, .2], #sets the drop out between cycles
                       emb_drop = .1  #embedding drop-out prior to training
                       )
#two cycles with 300 rows then 100 rows in the matrix
# ===========================================
dep_var = 'salary'
cat_names = ['workclass', 'education', 'marital-status', 'occupation', 'relationship', 'race']
cont_names = ['age', 'fnlwgt', 'education-num']
procs = [FillMissing, Categorify, Normalize]

test = TabularList.from_df(df.iloc[800:1000].copy(), path=path, cat_names=cat_names, cont_names=cont_names)

data = (TabularList.from_df(df, path=path, cat_names=cat_names, cont_names=cont_names, procs=procs)
                           .split_by_idx(list(range(800,1000)))
                           .label_from_df(cols=dep_var)
                           .add_test(test, label=0)
                           .databunch())
data.show_batch(rows=10)
learn = tabular_learner(data, layers=[200,100], metrics=accuracy)
learn.fit(1, 1e-2)
# ===========================================
# Collaborative filtering example
# collab models use data in a DataFrame of user, items, and rating
from fastai.collab import *
from fastai.tabular import *
user,item,title = 'userId','movieId','title'
data = CollabDataBunch.from_df(ratings, seed=42)
y_range = [0,5.5]
learn = collab_learner(data, n_factors=50, y_range=y_range)
learn.fit_one_cycle(3, 5e-3)
data = CollabDataBunch.from_df(rating_movie, seed=42, valid_pct=0.1, item_name=title)
learn = collab_learner(data, n_factors=40, y_range=y_range, wd=1e-1)
# ===========================================
# useful methods
download_images(urls, dest, max_pics)
verify_images(path, delete=True, max_workers=8)

np.random.seed(42)  # set to 42 to keep generate the same validation set
data=ImageDataBunch.from_folder(path, train='.', valid_pct=0.2,
        ds_tfms=get_transforms(), size=224, num_workers=4).normalized(imagenet_stats)
data.show_batch(row=2,figsize=(7.5))
data.lclasses, data.c, len(data.train_ds), len(data.valid_ds)

learn.freeze()          # only train the last few layers, keep the majority of info from previosu learner
learn.unfreeze()                # unfreeze all layers

learn.lr_find()                 # lr_find(learn)
learn.recorder.plot()           #plor learning rate
learn.recorder.plot_losses()    #plot training and validation loss
learn.recorder.plot_lr()        #plot lr
learn.fit_one_cycle(2,max_lr=slice(3e-5,3e-4))  # workflow 2
learn.save('stage-2')

interp = ClassificationInterpretation.from_learner(learn)
interp.plot_confusion_matrix()
       
doc()
??FileDeleter
ipywidget


# Normalize:
# A normilization takes continuous variables and subtract their mean and divide by their standard deviation so they are zero-one variable


# feather
# pip install feather-format


#================================================== 
from wordcloud import WordCloud
def grey_color_func(word, font_size, position, orientation, random_state=None, **kwargs):
    return "hsl(0, 0%%, %d%%)" % random.randint(60,100)
d = {}
for keyword, volume in df.values:
    d[keyword] = volume
wc= WordCloud(width=800,height=400)
plt.figure(figsize=(20,10))
wc.generate_from_frequencies(frequencies=d)
plt.imshow(wc.recolor(color_func=grey_color_func, random_state=3),interpolation='bilinear')
plt.axis('off')
plt.show
wc.to_file('myfile.jpg')

job_title= data['Job Title'][~pd.isnull(data['Job Title'])]
wordCloud=WorkCloud(width=450,height=300).generate(' '.join(job_title))
plt.figure(figsize=(19,9))
plt.axis('off')
plt.imshow(workCloud)
plt.show()

# include variable in label
plt.plot(x.lstat,x.ldf,'b-',label='posl %.2f'%c)


# return function annotation
def kinetic_energy(m:'in KG',v:'in M/S') ->'Joules':
    return 1/2*m*v**2
>> kinetic_energy.__annotations__
{'return': 'Joules','v':'in M/S','m':'in KG'}


# ====
# run python in command line 
import sys
def main():
    pass

if __name__ == "__main__":
    sys.exit(main())
# Parser for command-line options
import argparse 


help(WorkObj)


return NotImplemented

def isoformat(self):
    pass
__str__ = isoformat

#===============================
# 
try:
    ..
except KeyboardInterrupt:
    return 0 # successful exit
except ValueError:
    raise ValueError("{} is not a valie time format".format(time))
    return 2
except (StatisticsError, ZeroDivisionError, TypeError):
    return []
except Exception as e:
    if DEBUG or TESTRUN or verbose:
        print(e)
        raise e
    reurn 2 # unit command line syntax error
k




def convert_to_number(value: str) -> Union[float,str]:
    try:
        return float(value)
    except ValueError:
        return value.strip()
def drop:none(values: List[float]) -> List[float]:
    try:
        return [x for x in values if x is not None]
    except TypeError:
        return values


# unittest
import unittest
# run unittest in console 
if __name__ == '__main__':
    unittest.main()
# run unittest in Jupyter:
if __name__ == '__main__':
    unittest.main(argv=['first-arg-is-ignored'], exit=False)
